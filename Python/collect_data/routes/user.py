# -*- coding: utf-8 -*-
# @Time : 2019/11/11 20:28
# @Author : 尹傲雄
# @contact : yinaoxiong@gmail.com
# @Desc : 用户数据相关处理

import openpyxl as excel
from flask import request, Blueprint
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.exc import SQLAlchemyError

from . import generate_result, generate_validator
from .. import config
from ..models.base_model import db
from ..models.user import User
from ..utils import is_excel_end
from ..wraps import token_check, super_admin_required

user_bp = Blueprint('user', __name__, url_prefix=config.URL_Prefix + '/user')


@user_bp.route('/register', methods=['POST'])
@token_check
@super_admin_required
def register(*args, **kwargs):
    """
    用户注册接口，可用于批量导入也可用于单个注册
    """
    if request.method == 'POST':
        fail_ids = []
        data = request.get_json()
        schema = {
            'newUsers': {'type': 'list',
                         'schema': {'type': 'dict',
                                    'schema': {'iemi': {'type': 'string', 'maxlength': 17},
                                               'realName': {'type': 'string', 'maxlength': 32},
                                               'phoneNumber': {'type': 'string', 'maxlength': 11},
                                               'password': {'type': 'string', 'maxlength': 30}}}}
        }
        v = generate_validator(schema)
        if not v(data):  # 对请求数据格式进行校验
            return generate_result(1, data=v.errors)
        new_users = data['newUsers']
        for index, val in enumerate(new_users):
            try:
                user = User(**val)
                user.set_password(user.password)
                db.session.add(user)
                db.session.commit()
            except SQLAlchemyError:
                db.session.rollback()
                fail_ids.append(index)
        return generate_result(0, data={'fail_ids': fail_ids})


@user_bp.route('/import_user', methods=['POST'])
@token_check
@super_admin_required
def import_user(*args, **kwargs):
    """
    从文件导入新用户
    """
    try:
        file = request.files['file']
    except KeyError:
        return generate_result(1)
    try:
        table = excel.load_workbook(file, read_only=True)
    except InvalidFileException:
        return generate_result(2, '仅支持.xlsx格式的文件')
    if 'user' not in table.sheetnames:
        return generate_result(2, '不存在名称为user的sheet')
    sheet = table['user']
    fail_row = []
    for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=5, values_only=True)):
        if is_excel_end(row):
            break
        new_user = User(iemi=row[0], realName=row[1], phoneNumber=row[2], password=str(row[3]), permission=row[4])
        new_user.set_password(new_user.password)
        try:
            db.session.add(new_user)
            db.session.commit()
        except SQLAlchemyError:
            db.session.rollback()
            fail_row.append(index)
    if len(fail_row) != 0:
        return generate_result(0, '添加部分用户数据失败', {'fail_row': fail_row})
    return generate_result(0, '添加用户数据成功')


@user_bp.route('/login', methods=['POST'])
def login():
    """
    用户登录接口
    """
    if request.method == 'POST':
        data = request.get_json()
        schema = {
            'iemi': {'type': 'string', 'maxlength': 17},
            'password': {'type': 'string', 'maxlength': 30}
        }
        v = generate_validator(schema)
        if not v(data):
            return generate_result(1, data=v.errors)
        user = User.query.filter_by(iemi=data['iemi']).first()

        # 判断用户是否存在
        if user is None:
            return generate_result(2, '该用户不存在或密码错误')
        if not user.check_password(data['password']):
            return generate_result(2, '该用户不存在或密码错误')
        return generate_result(0, data={'token': str(user.generate_auth_token(), encoding='utf-8')})


@user_bp.route('/refresh_token', methods=['POST'])
@token_check
def refresh_token(user_id: int, *args, **kwargs):
    """
    更新token接口
    """
    user = User.query.get(user_id)
    return generate_result(0, data={'token': str(user.generate_auth_token(), encoding='utf-8')})


@user_bp.route('/all_info', methods=['POST'])
@token_check
@super_admin_required
def all_info(*args, **kwargs):
    """
    获取用户信息列表
    """
    user_info_list = [i.to_dict for i in User.query.all()]
    for user_info in user_info_list:
        del user_info['password']
    return generate_result(0, '获取用户列表成功', {'all_info': user_info_list})


@user_bp.route('/update', methods=['POST'])
@token_check
def update(user_id: int, *args, **kwargs):
    """
    用户更新个人信息
    :param user_id: 用户id
    """
    data = request.get_json()
    schema = {
        "realName": {'type': 'string', 'maxlength': 32},
        "phoneNumber": {'type': 'string'}
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(1, data=v.errors)
    user = User.query.get(user_id)
    user.realName = data['realName']
    user.phoneNumber = data['phoneNumber']
    try:
        db.session.add(user)
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        return generate_result(2, '更新信息失败')
    return generate_result(0, '更新用户信息成功')


@user_bp.route('/super_update', methods=['POST'])
@token_check
@super_admin_required
def super_update(*args, **kwargs):
    """
    超级管理员批量更新用户信息
    """
    data = request.get_json()
    schema = {
        'userInfo': {'type': 'list',
                     'schema': {'type': 'dict',
                                'schema': {'id': {'type': 'integer', 'min': 1},
                                           'realName': {'type': 'string', 'maxlength': 32},
                                           'phoneNumber': {'type': 'string', 'maxlength': 11},
                                           'password': {'type': 'string', 'maxlength': 30},
                                           'permission': {'type': 'integer', 'min': 0}}}}
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(1, data=v.errors)
    user_info = data['userInfo']
    fail_ids = []

    for index, val in enumerate(user_info):
        try:
            user = User.query.get(val['id'])
            user.update(**val)
            user.reset_password(user.password)
            db.session.add(user)
            db.session.commit()
        except SQLAlchemyError:
            db.session.rollback()
            # 回退redis
            User.redis_del(val['id'])
            fail_ids.append(index)
    return generate_result(0, data={'fail_ids': fail_ids})


@user_bp.route('/reset_password', methods=['POST'])
def reset_password():
    """
    用户更新密码接口
    :return:
    """
    data = request.get_json()
    schema = {
        'oldPassword': {'type': 'string', 'maxlength': 64},
        'newPassword': {'type': 'string', 'maxlength': 64},
        'iemi': {'type': 'string', 'maxlength': 17}
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(1)
    try:
        user = User.query.filter_by(iemi=data['iemi']).first()
    except SQLAlchemyError as e:
        print(str(e))
        db.session.rollback()
        return generate_result(2, '查询用户失败')
    if user is None:
        return generate_result(2, '用户不存在')

    try:
        if not user.check_password(data['oldPassword']):
            return generate_result(2, '密码错误')
        user.reset_password(data['newPassword'])
        db.session.add(user)
        db.session.commit()
    except SQLAlchemyError:
        db.session.rollback()
        # 回退redis
        User.redis_del(user.id)
        return generate_result(2, '数据更新失败')
    return generate_result(0, '更新密码成功')


@user_bp.route('/delete', methods=['POST'])
@token_check
@super_admin_required
def delete(*args, **kwargs):
    """
    删除用户
    :param user_id: 用户id
    :return:
    """
    data = request.get_json()
    schema = {
        'userId': {'type': 'integer', 'min': 1}
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(1, data=v.errors)

    the_user = User.query.get(data['userId'])
    if the_user is None:
        return generate_result(2, '用户不存在')
    try:
        db.session.delete(the_user)
        db.session.commit()
    except SQLAlchemyError as e:
        print(str(e))
        db.session.rollback()
        return generate_result(2, '用户以录入信息无法删除')
    return generate_result(0, '删除用户成功')
