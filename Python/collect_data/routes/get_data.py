# -*- coding: utf-8 -*-
# @Time : 2019/11/11 23:25
# @Author : 尹傲雄
# @contact : yinaoxiong@gmail.com
# @Desc : 数据获取模块

import os
from io import BytesIO

import openpyxl as excel
from flask import request, Blueprint, send_from_directory
from sqlalchemy.exc import SQLAlchemyError

from . import generate_validator, generate_result, my_send_file
from .. import config
from ..models.base_model import db
from ..models.building_import_info import BuildingImportInfo
from ..models.building_info import BuildingInfo
from ..models.building_picture import BuildingPicture
from ..models.city import City
from ..models.community import Community
from ..models.district import District
from ..models.garden import Garden
from ..models.garden_base_info import GardenBaseInfo
from ..models.garden_import_info import GardenImportInfo
from ..models.garden_picture import GardenPicture
from ..models.map_data import MapData
from ..models.other_picture import OtherPicture
from ..models.street import Street
from ..models.user import User
from ..utils import gcj02_to_bd09
from ..wraps import token_check, admin_required

get_data_bp = Blueprint('get_data', __name__, url_prefix=config.URL_Prefix + '/get_data')


def set_form_value(form, value):
    for key in value.keys():
        if value[key] is None:
            value[key] = ''
    for item in form:
        if item['type'] != 'list':
            if item['key'] != '' and item['key'] in value:
                item['value'] = value[item['key']]
                if item['type'] == 'multiple':
                    item['value'] = item['value'].split(',')
    return form


def generate_form(columns, forbid_column):
    result = []
    for col in columns:
        if col.name not in forbid_column:
            if col.type == db.Integer or col.type == db.Float:
                result.append({
                    'label': col.comment,
                    'key': col.name,
                    'required': False,
                    'changed': True,
                    'type': 'number',
                    'value': ''
                })
            else:
                result.append({
                    'label': col.comment,
                    'key': col.name,
                    'required': False,
                    'changed': True,
                    'type': 'text',
                    'value': ''
                })

    return result


@get_data_bp.route('/map', methods=['POST'])
@token_check
def map_data(*args, **kwargs):
    data = request.get_json()
    schema = {
        'gardenId': {'type': 'integer', 'min': 1},
        'mapId': {'type': 'integer', 'min': 0}
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(1)
    try:
        all_map_data = MapData.query.filter_by(gardenId=data['gardenId']).all()
    except SQLAlchemyError:
        return generate_result(2, '查询数据失败')
    if data['mapId'] == 1:
        for item in all_map_data:
            item.longitude, item.latitude = gcj02_to_bd09(item.longitude, item.latitude)
    return generate_result(0, '获取地图数据成功', {'map_data': [i.to_dict for i in all_map_data]})


@get_data_bp.route('/garden_base_info', methods=['POST'])
@token_check
def garden_base_info(*args, **kwargs):
    """
    获取小区基本信息
    """
    data = request.get_json()
    schema = {
        'gardenId': {'type': 'integer', 'min': 1}
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(1)
    garden = Garden.query.get(data['gardenId'])
    if garden is None:
        return generate_result(2, '小区不存在')

    city = City.query.get(garden.cityId)
    district = District.query.get(garden.districtId)
    street = Street.query.get(garden.streetId)
    community = Community.query.get(garden.communityId)
    garden_info = GardenBaseInfo.query.get(data['gardenId'])

    result = [
        {
            'label': '市',
            'key': '',
            'required': False,
            'changed': False,
            'type': 'text',
            'value': city.name
        },
        {
            'label': '区县',
            'key': '',
            'required': False,
            'changed': False,
            'type': 'text',
            'value': district.name
        },
        {
            'label': '街道',
            'key': '',
            'required': False,
            'changed': False,
            'type': 'text',
            'value': street.name
        },
        {
            'label': '社区名称',
            'key': '',
            'required': False,
            'changed': False,
            'type': 'text',
            'value': community.name
        },
        {
            'label': '社区别名',
            'key': 'communityAlias',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '小区名称',
            'key': '',
            'required': False,
            'changed': False,
            'type': 'text',
            'value': garden.name
        },
        {
            'label': '小区别名',
            'key': 'gardenAlias',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '小区别名2',
            'key': 'gardenAlias2',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '小区座落',
            'key': 'gardenLocation',
            'required': True,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '小区东至',
            'key': 'gardenEastTo',
            'required': True,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '小区西至',
            'key': 'gardenWestTo',
            'required': True,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '小区北至',
            'key': 'gardenNorthTo',
            'required': True,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '小区南至',
            'key': 'gardenSouthTo',
            'required': True,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '区域位置',
            'key': 'regionalLocation',
            'required': True,
            'changed': True,
            'type': 'radio',
            'option': ['城区中心', '城东区域', '城南区域', '城西区域', '城北区域', '近郊组团及乡镇', '远郊组团及乡镇'],
            'value': ''
        },
        {
            'label': '楼盘状态',
            'key': 'houseStatus',
            'required': True,
            'changed': True,
            'type': 'radio',
            'option': ["存量", "在建", "在售", "灭失"],
            'value': ''
        },
        {
            'label': '小区类型',
            'key': 'gardenKind',
            'required': True,
            'changed': True,
            'type': 'radio',
            'option': ["普通商品房", "老式住宅区", "经济适用房", "中档商品房", "高档商品房", "自建房", "零星住宅区", "农村拆迁房", "城市拆迁房", "人才专项房"],
            'value': ''
        },
        {
            "label": "建筑类型",
            "key": "buildingKind",
            "required": False,
            "changed": True,
            "type": "multiple",
            "option": ["住宅(电梯房)", "住宅(楼梯房)", "住宅(洋房)", "单身公寓(住宅)", "单身公寓(非住宅)", "办公写字楼", "别墅(独栋)", "别墅(联排)", "别墅(双拼)",
                       "叠墅", "自建民房", "其它类型"],
            "value": []
        },
        {
            'label': '房屋性质',
            'key': 'roomType',
            'required': True,
            'changed': True,
            'type': 'radio',
            'option': ["商品房", "房改房", "经济适用房", "集资房", "私房", "非商品房", "公租房", "安置房"],
            'value': ''
        },
        {
            'label': '建筑结构',
            'key': 'buildingStructure',
            'required': True,
            'changed': True,
            'type': 'radio',
            'option': ["钢混结构", "混合结构", "钢及钢混结构", "砖混结构", "砖木结构", "钢结构", "木结构", "简易结构"],
            'value': ''
        },
        {
            'label': '住宅幢数',
            'key': 'houseNumber',
            'required': True,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '非住宅幢数',
            'key': 'notHouseNumber',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '幢数描述',
            'key': 'description',
            'required': True,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '建成年份',
            'key': 'buildYear',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '设定年份',
            'key': 'setYear',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '土地性质',
            'key': 'landStatus',
            'required': True,
            'changed': True,
            'type': 'radio',
            'option': ["国有土地", "集体土地"],
            'value': ''
        },
        {
            'label': '使用权',
            'key': 'right',
            'required': True,
            'changed': True,
            'type': 'radio',
            'option': ["出让", "划拨", "集体"],
            'value': ''
        },
        {
            'label': '土地等级',
            'key': 'landGrade',
            'required': True,
            'changed': True,
            'type': 'radio',
            'option': ["一级", "二级", "三级", "四级", "五级", "六级", "七级", "八级", "九级", "十级", "十一级"],
            'value': ''
        },
        {
            'label': '询价记录',
            'key': 'askRecode',
            'required': True,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '是否封闭',
            'key': 'closed',
            'required': True,
            'changed': True,
            'type': 'radio',
            'option': ["是", "否"],
            'value': ''
        },
        {
            'label': '物管分类',
            'key': 'managementKind',
            'required': True,
            'changed': True,
            'type': 'radio',
            'option': ["专业物业管理", "社区保洁", "社区准物业", "无物业管理"],
            'value': ''
        },
        {
            'label': '价格初判',
            'key': 'beginPrice',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '小区概况表其它信息备注',
            'key': 'otherInfo',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '相邻小区',
            'key': 'neighborGarden',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '交通干道',
            'key': 'mainRoad',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            "label": "道路等级",
            "key": "roadGrade",
            "required": True,
            "changed": True,
            "type": "multiple",
            "option": ["主干道、快速路", "次干道", "街巷", "里弄", "特殊类型"],
            "value": []
        },
        {
            'label': '公交站名',
            'key': 'busStation',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '站点距离',
            'key': 'busStationDistance',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '普通公交',
            'key': 'baseBus',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '快速公交',
            'key': 'quickBus',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '线路条数',
            'key': 'busLines',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '地铁站',
            'key': 'subwayStation',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '地铁站距离',
            'key': 'subwayDistance',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '地铁线路条数',
            'key': 'subwayLines',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '农贸市场',
            'key': 'farmerMarket',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '超市商场',
            'key': 'market',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '医疗设施',
            'key': 'hospital',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '金融设施',
            'key': 'bank',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '文体设施',
            'key': 'gym',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '行政机关',
            'key': 'organization',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '幼儿园',
            'key': 'kindergarten',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '小学',
            'key': 'primary',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '中学',
            'key': 'middle',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '大学',
            'key': 'college',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '景点',
            'key': 'attractions',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '公园',
            'key': 'park',
            'required': False,
            'changed': True,
            'type': 'map',
            'radius': 100,
            'value': ''
        },
        {
            'label': '路牌号',
            'key': 'streetNumber',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '所处商圈',
            'key': 'businessArea',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '所处板块',
            'key': 'locationArea',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '周边利用',
            'key': 'aroundUse',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '河流山川',
            'key': 'riversAndMountains',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '噪音污染',
            'key': 'noisePollution',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '空气污染',
            'key': 'airPollution',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '不利设施',
            'key': 'adverseFacilities',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '其他污染',
            'key': 'otherPollution',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '其他影响',
            'key': 'otherInfluences',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '繁华程度',
            'key': 'busyDegree',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '休闲设施',
            'key': 'relaxFacilities',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '运动设施',
            'key': 'sportFacilities',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '安保设施',
            'key': 'securityFacilities',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '建筑风格',
            'key': 'architecturalStyle',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '小区绿化',
            'key': 'gardenGreening',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '小区评价',
            'key': 'gardenEvaluation',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '物业公司',
            'key': 'propertyCompany',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
    ]

    if garden_info is not None:
        result = set_form_value(result, garden_info.to_dict)

    return generate_result(0, '获取数据成功', {'gardenInfoList': result})


@get_data_bp.route('/building_base_info', methods=['POST'])
@token_check
def building_base_info(*args, **kwargs):
    """
    获取楼栋基本信息
    """
    result = [
        {
            'label': '楼幢名称',
            'key': 'buildingName',
            'required': True,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '楼幢别名',
            'key': 'buildingAlias',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '楼幢类别',
            'key': 'buildingKind',
            'required': True,
            'changed': True,
            'type': 'radio',
            "option": ["a 住宅(电梯房)", "b 住宅(楼梯房)", "c 住宅(洋房)", "d 单身公寓(住宅)", "e 单身公寓(非住宅)", "f 办公写字楼", "g 别墅(独栋)",
                       "h 别墅(联排)", "i 别墅(双拼)", "j 叠墅", "k 自建民房", "L 其它类型"],
            'value': ''
        },
        {
            'label': '是否是别墅',
            'key': 'isVilla',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ['是', '否'],
            'value': ''
        },
        {
            "label": "1F情况说明",
            "type": "list",
            "length": 6
        },
        {
            'label': '架空层',
            'key': 'overheadLayer',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ['_', '√'],
            'value': ''
        },
        {
            'label': '地上车库',
            'key': 'aboveGroundGarage',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ['_', '√'],
            'value': ''
        },
        {
            'label': '杂物间',
            'key': 'utilityRoom',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ['_', '√'],
            'value': ''
        },
        {
            'label': '住宅',
            'key': 'residential',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ['_', '√'],
            'value': ''
        },
        {
            'label': '商铺',
            'key': 'shop',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ['_', '√'],
            'value': ''
        },
        {
            'label': '一楼其它情况备注',
            'key': 'firstFloorOtherDescription',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            "label": "楼幢层数情况",
            "type": "list",
            "length": 12
        },
        {
            'label': '一单元',
            'key': 'oneUnitLayers',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '二单元',
            'key': 'twoUnitLayers',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '三单元',
            'key': 'threeUnitLayers',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '四单元',
            'key': 'fourUnitLayers',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '五单元',
            'key': 'fiveUnitLayers',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '六单元',
            'key': 'sixUnitLayers',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '七单元',
            'key': 'sevenUnitLayers',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '八单元',
            'key': 'eightUnitLayers',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '九单元',
            'key': 'nightUnitLayers',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '十单元',
            'key': 'tenUnitLayers',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '地上总层',
            'key': 'floorOverGround',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '地下总层',
            'key': 'floorUnderGround',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            "label": "物业分类",
            "key": "propertyKind",
            "required": False,
            "changed": True,
            "type": "radio",
            "option": ["低层", "多层", "多层_电梯房", "小高层", "高层", "超高层"],
            "value": ""
        },
        {
            "label": "建筑结构",
            "key": "buildingStructure",
            "required": False,
            "changed": True,
            "type": "radio",
            "option": ["钢混结构", "混合结构", "钢及钢混结构", "砖混结构", "砖木结构", "钢结构", "木结构", "简易结构"],
            "value": ""
        },
        {
            "label": "一梯户数情况",
            "type": "list",
            "length": 12
        },
        {
            'label': '一单元',
            'key': 'oneUnitHouseholds',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '二单元',
            'key': 'twoUnitHouseholds',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '三单元',
            'key': 'threeUnitHouseholds',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '四单元',
            'key': 'fourUnitHouseholds',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '五单元',
            'key': 'fiveUnitHouseholds',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '六单元',
            'key': 'sixUnitHouseholds',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '七单元',
            'key': 'sevenUnitHouseholds',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '八单元',
            'key': 'eightUnitHouseholds',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '九单元',
            'key': 'nightUnitHouseholds',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        }, {
            'label': '十单元',
            'key': 'tenUnitHouseholds',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '单元数',
            'key': 'numberOfUnit',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '一梯几户',
            'key': 'oneLiftNumber',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            "label": "户室号分布情况",
            "type": "list",
            "length": 8
        },
        {
            'label': '单元名称',
            'key': 'unitName',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '室号名称',
            'key': 'roomName',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '单元号',
            'key': 'unitNumber',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '楼层号',
            'key': 'floorNumber',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '楼层差异',
            'key': 'floorDifferent',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '住宅起始',
            'key': 'beginFloor',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '店铺层数',
            'key': 'shopLayer',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '画板',
            'key': 'palette',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '部位说明',
            'key': 'locationDescription',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            'label': '主要朝向',
            'key': 'mainTowards',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ["东", "南", "西", "北", "东南", "东北", "西南", "西北"],
            'value': ''
        },
        {
            'label': '平面布局',
            'key': 'planeLayout',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ["一般", "较好", "好", "较差", "差"],
            'value': ''
        },
        {
            'label': '建成年份',
            'key': 'completedYear',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '装修标准',
            'key': 'decorationStandard',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ["毛坯房", "简单装修", "中档装修", "高档装修", "豪华装修"],
            'value': ''
        },
        {
            'label': '装修时间',
            'key': 'decorationYear',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '房产性质',
            'key': 'buildingProperty',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ["商品房", "房改房", "经济适用房", "集资房", "私房", "非商品房", "公租房", "安置房"],
            'value': ''
        },
        {
            'label': '楼幢状态',
            'key': 'buildingStatus',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ['存量', '灭失'],
            'value': ''
        },
        {
            "label": "顶楼状况",
            "type": "list",
            "length": 5
        },
        {
            'label': '顶楼情况',
            'key': 'roofTopInfo',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ["平层", "跃层", "阁楼", "露台"],
            'value': ''
        },
        {
            'label': '屋面情况',
            'key': 'roofInfo',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ['坡屋顶', '平屋顶'],
            'value': ''
        },
        {
            'label': '顶楼露台',
            'key': 'roofTopTerrace',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ['有', '无'],
            'value': ''
        },
        {
            'label': '顶楼阁楼',
            'key': 'roofTopAttic',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ['有', '无'],
            'value': ''
        },
        {
            'label': '顶楼跃层',
            'key': 'roofTopLayer',
            'required': False,
            'changed': True,
            'type': 'radio',
            "option": ['有', '无'],
            'value': ''
        },
        {
            'label': '楼栋调查表其他备注信息',
            'key': 'otherInfo',
            'required': False,
            'changed': True,
            'type': 'text',
            'value': ''
        },
        {
            "label": " 楼幢综合评价打分",
            "type": "list",
            "length": 8
        },
        {
            'label': '小区外视野景观',
            'key': 'outsideViewScore',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '中心花园位置',
            'key': 'centralGardenScore',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '通风采光情况',
            'key': 'ventilatedLightingScore',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '三临情况',
            'key': 'threeAdventSituationScore',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '周围环境',
            'key': 'surroundingsScore',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '建筑外观',
            'key': 'buildingAppearanceScore',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '其他有利',
            'key': 'otherAdvantagesScore',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '其他不利',
            'key': 'otherDisadvantagesScore',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '楼幢标准层等级初判',
            'key': 'buildingLevel',
            'required': False,
            'changed': True,
            'type': 'number',
            'value': ''
        },
        {
            'label': '价格初判',
            'key': 'unitNumber',
            'required': False,
            'changed': True,
            'type': 'preliminaryPrice',
            'value': ''
        },
    ]
    data = request.get_json()
    schema = {
        'buildingId': {'type': 'integer', 'min': 1}
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(0, data={'buildingInfoList': result})
    try:
        building_info = BuildingInfo.query.get(data['buildingId'])
    except SQLAlchemyError:
        return generate_result(2, '获取数据失败')
    if building_info is None:
        return generate_result(2, '楼栋不存在')
    building_info = building_info.to_dict
    result = set_form_value(result, building_info)

    return generate_result(0, '获取数据成功', {'buildingInfoList': result})


@get_data_bp.route('/garden_import_info', methods=['POST'])
@token_check
def garden_import_info(*args, **kwargs):
    """
    获取小区导入数据接口
    """
    data = request.get_json()
    schema = {
        'gardenId': {'type': 'integer', 'min': 1}
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(1)
    garden = Garden.query.get(data['gardenId'])
    if garden is None:
        return generate_result(2, '小区不存在')
    result = [
        {
            'label': '小区名称',
            'key': '',
            'required': False,
            'changed': False,
            'type': 'text',
            'value': garden.name
        },
    ]

    result.extend(GardenImportInfo().generate_form(['id', 'userId', 'collectTime']))

    import_info = GardenImportInfo.query.get(data['gardenId'])

    if import_info is not None:
        result = set_form_value(result, import_info.to_dict)
    return generate_result(0, '获取数据成功', {'gardenImportInfoList': result})


@get_data_bp.route('/building_import_info', methods=['POST'])
@token_check
def building_import_info(*args, **kwargs):
    data = request.get_json()
    schema = {
        'buildingId': {'type': 'integer', 'min': 1}
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(1)
    building = BuildingInfo.query.get(data['buildingId'])
    if building is None:
        return generate_result(2, '楼幢不存在')
    result = [
        {
            'label': '楼幢名称',
            'key': '',
            'required': False,
            'changed': False,
            'type': 'text',
            'value': building.buildingName
        }
    ]
    result.extend(BuildingImportInfo().generate_form(['id', 'userId', 'collectTime']))

    import_info = BuildingImportInfo.query.get(data['buildingId'])

    if import_info is not None:
        result = set_form_value(result, import_info.to_dict)

    return generate_result(0, '获取楼幢数据成功', data={'buildingImportInfoList': result})


@get_data_bp.route('/garden_base_table', methods=['POST'])
@token_check
@admin_required
def garden_base_table(*args, **kwargs):
    """
    导出表1 《小区概况表》
    """
    try:
        garden_id = request.form['gardenId']
    except KeyError:
        return generate_result(1)
    try:
        garden = Garden.query.get(garden_id)
        if garden is None:
            return generate_result(2, '小区不存在')
        city = City.query.get(garden.cityId)
        district = District.query.get(garden.districtId)
        street = Street.query.get(garden.streetId)
        community = Community.query.get(garden.communityId)
        garden_info = GardenBaseInfo.query.get(garden_id)
        user = User.query.get(garden_info.userId)
    except SQLAlchemyError:
        return generate_result(2, '导出数据失败')
    if garden_info is None:
        return generate_result(2, '小区基本信息不存在')
    father_dictionary = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
    file_path = os.path.join(father_dictionary, 'data_template', '小区概况表导出模板.xlsx')
    wb = excel.load_workbook(file_path)
    ws = wb.active
    ws.title = '表1 《小区概况表》'
    table_map = {'采集人员': 'userName', '采集日期': 'collectTime', '市': 'city', '区县': 'district', '街道': 'street',
                 '社区名称': 'community', '社区别名': 'communityAlias', '小区名称': 'gardenName', '小区别名': 'gardenAlias',
                 '小区别名2': 'gardenAlias2', '小区座落': 'gardenLocation', '小区东至': 'gardenEastTo', '小区南至': 'gardenWestTo',
                 '小区西至': 'gardenNorthTo', '小区北至': 'gardenSouthTo', '区域位置': 'regionalLocation', '楼盘状态': 'houseStatus',
                 '小区类型': 'gardenKind', '建筑类型': 'buildingKind', '房屋性质': 'roomType', '建筑结构': 'buildingStructure',
                 '住宅幢数': 'houseNumber', '非住宅幢数': 'notHouseNumber', '幢数描述': 'description', '建成年份': 'buildYear',
                 '设定年份': 'setYear', '土地性质': 'landStatus', '使_用_权': 'right', '土地等级': 'landGrade', '询价记录': 'askRecode',
                 '是否封闭': 'closed', '物管分类': 'managementKind', '价格初判': 'beginPrice', '小区概况表其它信息备注': 'otherInfo',
                 '相邻小区': 'neighborGarden', '交通干道': 'mainRoad', '道路等级': 'roadGrade', '公交站名': 'busStation',
                 '站点距离': 'busStationDistance', '普通公交': 'baseBus', '快速公交': 'quickBus', '线路条数': 'busLines',
                 '地铁站名': 'subwayStation', '地铁距离': 'subwayDistance', '地铁线路': 'subwayLines', '农贸市场': 'farmerMarket',
                 '超市商场': 'market', '医疗设施': 'hospital', '金融机构': 'bank', '文体场馆': 'gym', '行政机关': 'organization',
                 '幼_儿_园': 'kindergarten', '小学教育': 'primary', '中学教育': 'middle', '大学教育': 'college', '旅游景点': 'attractions',
                 '公园广场': 'park', '路_牌_号': 'streetNumber', '所处商圈': 'businessArea', '所处板块': 'locationArea',
                 '周边利用': 'aroundUse', '河流山脉': 'riversAndMountains', '噪音污染': 'noisePollution', '空气污染': 'airPollution',
                 '不利设施': 'adverseFacilities', '其他污染': 'otherPollution', '其他影响': 'otherInfluences', '繁华程度': 'busyDegree',
                 '休闲设施': 'relaxFacilities', '运动设施': 'sportFacilities', '安保设施': 'securityFacilities',
                 '建筑风格': 'architecturalStyle', '小区绿化': 'gardenGreening', '小区评价': 'gardenEvaluation',
                 '物管公司': 'propertyCompany'}
    garden_info = garden_info.to_dict
    garden_info['userName'] = user.realName
    garden_info['gardenName'] = garden.name
    garden_info['city'] = city.name
    garden_info['district'] = district.name
    garden_info['street'] = street.name
    garden_info['community'] = community.name
    # 处理None值的情况
    for key in garden_info.keys():
        if garden_info[key] is None:
            garden_info[key] = ''

    for row in ws.iter_rows(min_row=1, max_row=75, max_col=2):
        key = table_map[row[0].value]
        row[1].value = garden_info[key]
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return my_send_file(stream, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', '小区概况表.xlsx')


@get_data_bp.route('/building_base_table', methods=['POST'])
def building_base_table(*args, **kwargs):
    """
    导出表2 《楼幢调查表》
    """
    try:
        garden_id = request.form['gardenId']
    except KeyError:
        return generate_result(1)
    try:
        building_infos = BuildingInfo.query.filter_by(gardenId=garden_id).all()
    except SQLAlchemyError as e:
        print(str(e))
        return generate_result(2, '导出数据失败')
    if len(building_infos) == 0:
        return generate_result(2, '楼幢数据不存在')

    # 获取模板文件路径
    father_dictionary = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
    file_path = os.path.join(father_dictionary, 'data_template', '楼幢调查表导出模板.xlsx')
    table_key_list = ['buildingName', 'buildingAlias', 'buildingKind', 'isVilla', 'overheadLayer', 'aboveGroundGarage',
                      'utilityRoom', 'residential', 'shop', 'firstFloorOtherDescription', 'oneUnitLayers',
                      'twoUnitLayers', 'threeUnitLayers', 'fourUnitLayers', 'fiveUnitLayers', 'sixUnitLayers',
                      'sevenUnitLayers', 'eightUnitLayers', 'nightUnitLayers', 'tenUnitLayers', 'floorOverGround',
                      'floorUnderGround', 'propertyKind', 'buildingStructure', 'oneUnitHouseholds', 'twoUnitHouseholds',
                      'threeUnitHouseholds', 'fourUnitHouseholds', 'fiveUnitHouseholds', 'sixUnitHouseholds',
                      'sevenUnitHouseholds', 'eightUnitHouseholds', 'nightUnitHouseholds', 'tenUnitHouseholds',
                      'numberOfUnit', 'oneLiftNumber', 'unitName', 'roomName', 'unitNumber', 'floorNumber',
                      'floorDifferent', 'beginFloor', 'shopLayer', 'palette', 'locationDescription', 'mainTowards',
                      'planeLayout', 'completedYear', 'decorationStandard', 'decorationYear', 'buildingProperty',
                      'buildingStatus', 'roofTopInfo', 'roofInfo', 'roofTopTerrace', 'roofTopAttic', 'roofTopLayer',
                      'otherInfo', 'outsideViewScore', 'centralGardenScore', 'ventilatedLightingScore',
                      'threeAdventSituationScore', 'surroundingsScore', 'buildingAppearanceScore',
                      'otherAdvantagesScore', 'otherDisadvantagesScore', 'buildingLevel', 'preliminaryPrice']
    wb = excel.load_workbook(file_path)
    ws = wb.active

    for item in building_infos:
        item = item.to_dict
        for key in item.keys():
            if item[key] is None:
                item[key] = ''
        ws.append([item[table_key] for table_key in table_key_list])

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return my_send_file(stream, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', '楼幢调查表.xlsx')


@get_data_bp.route('/garden_table', methods=['POST'])
@token_check
@admin_required
def garden_table(*args, **kwargs):
    """
    导出小区信息_数据导入表
    """
    try:
        garden_id = request.form['gardenId']
    except KeyError:
        return generate_result(1)
    try:
        garden = Garden.query.get(garden_id)
        if garden is None:
            return generate_result(2, '小区不存在')
        garden_info = GardenBaseInfo.query.get(garden_id)
        if garden_info is None:
            return generate_result(2, '小区基本信息不存在')
        import_info = GardenImportInfo.query.get(garden_id)
    except SQLAlchemyError:
        return generate_result(2, '导出数据失败')

    # 获取模板文件路径
    father_dictionary = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
    file_path = os.path.join(father_dictionary, 'data_template', '小区信息_数据导入表导出模板.xlsx')
    wb = excel.load_workbook(file_path)
    ws = wb.active
    ws.title = '表3 《小区信息_数据导入表》'
    table_key_list = ['gardenName', 'gardenAlias', 'streetNumber', 'boundaryStreetSign', 'gardenEastTo',
                      'gardenSouthTo', 'gardenWestTo', 'gardenNorthTo', 'regionalLocation', 'businessArea',
                      'locationArea', 'houseStatus', 'gardenKind', 'buildingKind', 'toomType', 'buildingStructure',
                      'buildingNumbers', 'buildYear', 'setYear', 'constructionUnit', 'volumeRate', 'buildingDensity',
                      'developer', 'greeningRate', 'occupancyRate', 'landStatus', 'right', 'landGrade', 'landArea',
                      'constructionScale', 'residentialArea', 'houseNumber', 'description', 'propertyCompany',
                      'salesAgent', 'salesAddress', 'tuQiu', 'salesPhone', 'salesTime', 'mapSource',
                      'positioningCoordinates', 'projectDescription', 'outEnvironment', 'neighborGarden', 'mainRoad',
                      'roadGrade', 'busStation', 'busStationDistance', 'baseBus', 'quickBus', 'busLines',
                      'subwayStation', 'subwayDistance', 'subwayLines', 'aroundUse', 'farmerMarket', 'market',
                      'hospital', 'bank', 'gym', 'organization', 'kindergarten', 'primary', 'middle', 'college',
                      'attractions', 'riversAndMountains', 'noisePollution', 'noisePollution', 'adverseFacilities',
                      'otherPollution', 'otherEffects', 'busyDegree', 'park', 'businessDistrictDistance',
                      'infrastructure', 'insideRoad', 'zoneEnvironment', 'gardenClubhouse', 'relaxFacilities',
                      'sportFacilities', 'numberOfBerths', 'berthType', 'berthRatio', 'businessPackage', 'closed',
                      'managementKind', 'propertyCosts', 'securityFacilities', 'architecturalStyle', 'gardenGreening',
                      'gardenEvaluation', 'beginPrice']

    garden_info = garden_info.to_dict
    # 添加小区名字
    garden_info['gardenName'] = garden.name
    if import_info is not None:
        import_info = import_info.to_dict
        # 合并dict
        all_garden_info = {**garden_info, **import_info}
    else:
        all_garden_info = garden_info
    # 对数据为none的情况进行处理
    for key in all_garden_info.keys():
        if all_garden_info[key] is None:
            all_garden_info[key] = ''
    table_data = []
    for table_key in table_key_list:
        if table_key not in all_garden_info:
            table_data.append('')
        else:
            table_data.append(all_garden_info[table_key])
    ws.append(table_data)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return my_send_file(stream, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', '小区信息_数据导入表.xlsx')


@get_data_bp.route('/building_table', methods=['POST'])
@token_check
@admin_required
def building_table(*args, **kwargs):
    """
    导出楼幢信息_数据导入表
    """
    try:
        garden_id = request.form['gardenId']
    except KeyError:
        return generate_result(1)
    try:
        building_info = BuildingInfo.query.filter_by(gardenId=garden_id).all()
        if len(building_info) == 0:
            return generate_result(2, '该小区暂无楼幢信息')
    except SQLAlchemyError:
        return generate_result(2, '导出数据失败')
    # 获取模板文件路径
    father_dictionary = os.path.abspath(os.path.dirname(os.path.dirname(__file__)))
    file_path = os.path.join(father_dictionary, 'data_template', '楼幢信息_数据导入表导出模板.xlsx')
    wb = excel.load_workbook(file_path)
    ws = wb.active
    ws.title = '表4《楼幢信息_数据导入表》'
    table_key_list = ["buildingName", "buildingAlias", "buildingEastTo", "buildingWestTo", "buildingSouthTo",
                      "buildingNorthTo", "propertyKind", "unitName", "roomName", "numberOfUnit", "unitNumber",
                      "floorNumber", "floorDifferent", "beginFloor", "allBuildingNumbers", "mainTowards",
                      "locationDescription", "buildingStructure", "completedYear", "floorOverGround",
                      "floorUnderGround", "housingHeight", "decorationStandard", "decorationDescription",
                      "decorationYear", "houseType", "buildingProperty", "tuQiu", "auxiliaryRoomUse", "overheadLayer",
                      "aboveGroundGarage", "underGroundGarage", "constructionArea", "isVilla", "oneLiftNumber",
                      "haveRoomRate", "buildingStatus", "architecturalStyle", "equipmentAndFacilities",
                      "exteriorWallFinishes", "fireFacilities", "elevatorFacilities", "airConditioningFacilities",
                      "hotWater", "lobbyEntrance", "roofTopInfo", "roofTopTerrace", "roofTopAttic", "roofTopLayer",
                      "roofInfo", "elevatorLobby", "publicParts", "planeLayout", "ventilatedLighting", "businessRoom",
                      "usage", "outsideView", "centralGarden", "unfavorableFacilities", "soundPolution",
                      "buildingSpacing", "threeAdventSituation", "otherDisadvantages", "otherAdvantages",
                      "buildingEvaluation", "preliminaryPrice", "mapSource", "positioningCoordinates"]
    for info in building_info:
        info = info.to_dict
        try:
            import_info = BuildingImportInfo.query.get(info['id'])
        except SQLAlchemyError as e:
            print(str(e))
            continue
        if import_info is not None:
            all_info = {**info, **import_info.to_dict}
        else:
            all_info = info
        for key in all_info.keys():
            if all_info[key] is None:
                all_info[key] = ''
        row_data = []
        for table_key in table_key_list:
            if table_key not in all_info:
                row_data.append('')
            else:
                row_data.append(all_info[table_key])
        ws.append(row_data)
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return my_send_file(stream, 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', '楼幢信息_数据导入表.xlsx')


@get_data_bp.route('/picture', methods=['GET'])
@token_check
@admin_required
def picture(*args, **kwargs):
    """
    根据图片路径获取图片
    :return:
    """
    relative_path = request.args.get('path', '')
    if relative_path == '':
        return generate_result(1)
    abs_path = os.path.join(config.UPLOADED_IMAGES_DEST, relative_path)
    if not os.path.exists(abs_path):
        return generate_result(2, '图片不存在')
    dir_path, filename = os.path.split(abs_path)
    return send_from_directory(dir_path, filename)


@get_data_bp.route('/garden_picture', methods=['POST'])
@token_check
@admin_required
def garden_picture(*args, **kwargs):
    """
    获取小区图片的相对路径
    """
    data = request.get_json()
    schema = {
        'id': {'type': 'integer', 'min': 1},
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(1)
    try:
        garden_pictures = GardenPicture.query.filter_by(gardenId=data['id']).all()
    except SQLAlchemyError as e:
        print(str(e))
        return generate_result(2)
    if len(garden_pictures) == 0:
        return generate_result(2, '该小区暂未上传图片')
    return generate_result(0, '获取小区图片成功', data={"gardenPictures": [i.to_dict for i in garden_pictures]})


@get_data_bp.route('/other_picture', methods=['POST'])
@token_check
@admin_required
def other_picture(*args, **kwargs):
    """
    获取小区其他图片
    """
    data = request.get_json()
    schema = {
        'id': {'type': 'integer', 'min': 1},
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(1)
    try:
        other_pictures = OtherPicture.query.filter_by(gardenId=data['id']).all()
    except SQLAlchemyError as e:
        print(str(e))
        return generate_result(2)
    if len(other_pictures) == 0:
        return generate_result(2, '该小区暂未上传图片')
    return generate_result(0, '获取小区图片成功', data={'otherPictures': [i.to_dict for i in other_pictures]})


@get_data_bp.route('/building_picture', methods=['POST'])
@token_check
@admin_required
def building_picture(*args, **kwargs):
    """
    获取楼幢图片接口
    """
    data = request.get_json()
    schema = {
        'id': {'type': 'integer', 'min': 1},
    }
    v = generate_validator(schema)
    if not v(data):
        return generate_result(1)
    try:
        building_pictures = BuildingPicture.query.filter_by(buildingId=data['id']).all()
    except SQLAlchemyError as e:
        print(str(e))
        return generate_result(2)
    if len(building_pictures) == 0:
        return generate_result(2, '该楼幢暂无数据')
    return generate_result(0, '获取楼幢图片成功', data={'buildingPictures': [i.to_dict for i in building_pictures]})


@get_data_bp.route('/disk', methods=['POST'])
@token_check
@admin_required
def disk(*args, **kwargs):
    import psutil
    disk_usage = psutil.disk_usage(os.getcwd())
    return generate_result(0, '获取磁盘数据成功',
                           data={'total': disk_usage.total, 'used': disk_usage.used, 'free': disk_usage.free,
                                 'percent': disk_usage.percent})
