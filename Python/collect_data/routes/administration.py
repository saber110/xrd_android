# -*- coding: utf-8 -*-
# @Time : 2019/11/11 20:09
# @Author : 尹傲雄
# @contact : yinaoxiong@gmail.com
# @Desc : 行政区数据获取

import openpyxl as excel
from flask import request, Blueprint
from openpyxl.utils.exceptions import InvalidFileException
from sqlalchemy.exc import SQLAlchemyError

from . import generate_result
from .. import config
from ..models.base_model import db
from ..models.city import City
from ..models.community import Community
from ..models.district import District
from ..models.garden import Garden
from ..models.province import Province
from ..models.street import Street
from ..utils import is_excel_end
from ..wraps import token_check, super_admin_required

administration_bp = Blueprint('administration', __name__, url_prefix=config.URL_Prefix + '/administration')


def add_province(province_name):
    the_province = Province.query.filter_by(name=province_name).first()
    if the_province is None:
        the_province = Province(name=province_name)
        try:
            db.session.add(the_province)
            db.session.commit()
        except SQLAlchemyError as e:
            print(str(e))
            db.session.rollback()
            return None
    return {'province': the_province}


def add_city(province_name, city_name):
    result = add_province(province_name)
    if result is None:
        return None
    the_province = result['province']
    the_city = City.query.filter_by(provinceId=the_province.id, name=city_name).first()
    if the_city is None:
        the_city = City(provinceId=the_province.id, name=city_name)
        try:
            db.session.add(the_city)
            db.session.commit()
        except SQLAlchemyError as e:
            print(str(e))
            db.session.rollback()
            return None
    result['city'] = the_city
    return result


def add_district(province_name, city_name, district_name):
    result = add_city(province_name, city_name)
    if result is None:
        return None
    the_city = result['city']
    the_district = District.query.filter_by(cityId=the_city.id, name=district_name).first()
    if the_district is None:
        the_district = District(cityId=the_city.id, name=district_name)
        try:
            db.session.add(the_district)
            db.session.commit()
        except SQLAlchemyError as e:
            print(str(e))
            db.session.rollback()
            return None
    result['district'] = the_district
    return result


def add_street(province_name, city_name, district_name, street_name):
    result = add_district(province_name, city_name, district_name)
    if result is None:
        return None
    the_district = result['district']
    the_street = Street.query.filter_by(districtId=the_district.id, name=street_name).first()
    if the_street is None:
        the_street = Street(districtId=the_district.id, name=street_name)
        try:
            db.session.add(the_street)
            db.session.commit()
        except SQLAlchemyError as e:
            print(str(e))
            db.session.rollback()
            return None
    result['street'] = the_street
    return result


def add_community(province_name, city_name, district_name, street_name, community_name):
    result = add_street(province_name, city_name, district_name, street_name)
    if result is None:
        return None
    the_street = result['street']
    the_community = Community.query.filter_by(streetId=the_street.id, name=community_name).first()
    if the_community is None:
        the_community = Community(streetId=the_street.id, name=community_name)
        try:
            db.session.add(the_community)
            db.session.commit()
        except SQLAlchemyError as e:
            print(str(e))
            db.session.rollback()
            return None
    result['community'] = the_community
    return result


def add_garden(province_name, city_name, district_name, street_name, community_name, garden_name):
    result = add_community(province_name, city_name, district_name, street_name, community_name)
    if result is None:
        return None
    the_community = result['community']
    the_street = result['street']
    the_district = result['district']
    the_city = result['city']
    the_province = result['province']
    the_garden = Garden.query.filter_by(communityId=the_community.id, streetId=the_street.id,
                                        districtId=the_district.id,
                                        cityId=the_city.id, provinceId=the_province.id, name=garden_name).first()
    if the_garden is None:
        the_garden = Garden(communityId=the_community.id, streetId=the_street.id, districtId=the_district.id,
                            cityId=the_city.id, provinceId=the_province.id, name=garden_name)
        try:
            db.session.add(the_garden)
            db.session.commit()
        except SQLAlchemyError as e:
            print(str(e))
            db.session.rollback()
            return None
    result['garden'] = the_garden
    return result


@administration_bp.route('/province', methods=['POST'])
@token_check
def province(*args, **kwargs):
    provinces = Province.query.all()
    data = {'provinces': [i.to_dict for i in provinces]}
    return generate_result(0, '获取省份数据成功', data)


@administration_bp.route('/city', methods=['POST'])
@token_check
def city(*args, **kwargs):
    data = request.get_json()
    if 'provinceId' in data:
        provinceId = data['provinceId']
        cities = City.query.filter_by(provinceId=provinceId).all()
        result_list = []
        for i in cities:
            the_dict = i.to_dict
            the_dict.pop('provinceId')
            result_list.append(the_dict)
        return generate_result(0, '获取城市数据成功', {'cities': result_list})
    return generate_result(1)


@administration_bp.route('/district', methods=['POST'])
@token_check
def district(*args, **kwargs):
    data = request.get_json()
    if 'cityId' in data:
        cityId = data['cityId']
        districts = District.query.filter_by(cityId=cityId).all()
        result_list = []
        for i in districts:
            the_dict = i.to_dict
            the_dict.pop('cityId')
            result_list.append(the_dict)
        return generate_result(0, '获取行政区数据成功', {'districts': result_list})
    return generate_result(1)


@administration_bp.route('/street', methods=['POST'])
@token_check
def street(*args, **kwargs):
    data = request.get_json()
    if 'districtId' in data:
        districtId = data['districtId']
        streets = Street.query.filter_by(districtId=districtId).all()
        result_list = []
        for i in streets:
            the_dict = i.to_dict
            the_dict.pop('districtId')
            result_list.append(the_dict)
        return generate_result(0, '获取街道数据成功', {'streets': result_list})
    return generate_result(1)


@administration_bp.route('/community', methods=['POST'])
@token_check
def community(*args, **kwargs):
    data = request.get_json()
    if 'streetId' in data:
        streetId = data['streetId']
        communities = Community.query.filter_by(streetId=streetId).all()
        result_list = []
        for i in communities:
            the_dict = i.to_dict
            the_dict.pop('streetId')
            result_list.append(the_dict)
        return generate_result(0, '获取社区数据成功', {'communities': result_list})
    return generate_result(1)


@administration_bp.route('/garden', methods=['POST'])
@token_check
def garden(*args, **kwargs):
    data = request.get_json()
    if 'communityId' in data:
        community_id = data['communityId']
        gardens = Garden.query.filter_by(communityId=community_id).all()
        result_list = []
        for i in gardens:
            the_dict = i.to_dict
            the_dict.pop('communityId')
            result_list.append(the_dict)
        return generate_result(0, '获取小区数据成功', {'gardens': result_list})
    return generate_result(1)


@administration_bp.route('/search_garden', methods=['POST'])
@token_check
def search_garden(*args, **kwargs):
    data = request.get_json()
    if 'gardenName' not in data:
        return generate_result(1)
    try:
        gardens = Garden.query.filter(Garden.name.like('%' + data['gardenName'] + '%')).all()
    except SQLAlchemyError as e:
        print(str(e))
        return generate_result(2, '搜索失败')
    if len(gardens) == 0:
        return generate_result(2, '小区不存在')
    result = []
    for garden in gardens:
        community = Community.query.get(garden.communityId)
        street = Street.query.get(garden.streetId)
        district = District.query.get(garden.districtId)
        city = City.query.get(garden.cityId)
        province = Province.query.get(garden.provinceId)
        result.append({
            'gardenId': garden.id,
            'gardenName': garden.name,
            'communityName': community.name,
            'streetName': street.name,
            'districtName': district.name,
            'cityName': city.name,
            'provinceName': province.name
        })
    return generate_result(0, '查找小区成功', data={'gardens': result})


@administration_bp.route('/import_data', methods=['POST'])
@token_check
@super_admin_required
def import_data(*args, **kwargs):
    """
    从excel文件导入行政区数据
    :return:
    """
    try:
        file = request.files['file']
    except KeyError:
        return generate_result(1)
    try:
        table = excel.load_workbook(file, read_only=True)
    except InvalidFileException:
        return generate_result(2, '仅支持.xlsx格式的文件')
    sheet_name_list = table.sheetnames
    fail_dict = {}
    if '省份' in sheet_name_list:
        sheet = table['省份']
        district_fail_row = []
        for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=1, values_only=True)):
            if is_excel_end(row):
                break  # 以防出现excel末尾存在大量空值
            if add_province(*row) is None:
                district_fail_row.append(index + 1)
        fail_dict['省份'] = district_fail_row
    if '省份_城市' in sheet_name_list:
        sheet = table['省份_城市']
        street_fail_row = []
        for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=2, values_only=True)):
            if is_excel_end(row):
                break
            if add_city(*row) is None:
                street_fail_row.append(index + 1)
        fail_dict['省份_城市'] = street_fail_row
    if '省份_城市_行政区' in sheet_name_list:
        sheet = table['省份_城市_行政区']
        district_fail_row = []
        for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=3, values_only=True)):
            if is_excel_end(row):
                break  # 以防出现excel末尾存在大量空值
            if add_district(*row) is None:
                district_fail_row.append(index + 1)
        fail_dict['省份_城市_行政区'] = district_fail_row
    if '省份_城市_行政区_街道' in sheet_name_list:
        sheet = table['省份_城市_行政区_街道']
        street_fail_row = []
        for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=4, values_only=True)):
            if is_excel_end(row):
                break
            if add_street(*row) is None:
                street_fail_row.append(index + 1)
        fail_dict['省份_城市_行政区_街道'] = street_fail_row
    if '省份_城市_行政区_街道_社区' in sheet_name_list:
        sheet = table['省份_城市_行政区_街道_社区']
        community_fail_row = []
        for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=5, values_only=True)):
            if is_excel_end(row):
                break
            if add_community(*row) is None:
                community_fail_row.append(index + 1)
        fail_dict['省份_城市_行政区_街道_社区'] = community_fail_row
    if '省份_城市_行政区_街道_社区_小区' in sheet_name_list:
        sheet = table['省份_城市_行政区_街道_社区_小区']
        garden_fail_row = []
        for index, row in enumerate(sheet.iter_rows(min_row=2, max_col=6, values_only=True)):
            if is_excel_end(row):
                break
            if add_garden(*row) is None:
                garden_fail_row.append(index + 1)
        fail_dict['省份_城市_行政区_街道_社区_小区'] = garden_fail_row

    return generate_result(0, '导入数据成功', {'fail_rows': fail_dict})
